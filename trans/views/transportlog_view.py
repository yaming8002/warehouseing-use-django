import sys

from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render
from django.views import View
from django.views.generic.edit import CreateView
from django.views.generic.list import ListView
from openpyxl import load_workbook

from report.models.steel_model import DoneSteelReport
from stock.models.material import MatCat, Materials
from stock.models.site import SiteInfo
from trans.models import  TransLog, TransLogDetail
from wcommon.utils.excel_tool import ImportDataGeneric
from wcommon.utils.pagelist import PageListView
from wcommon.utils.uitls import excel_value_to_str, get_month_range
import logging
# # Create your models here.
import logging.config
from django.conf import settings

logging.config.dictConfig(settings.LOGGING)

logger = logging.getLogger(__name__)

class TrandportView(PageListView):
    model = TransLogDetail
    template_name = "trans/transport_log.html"
    title_name = "出入總表"

    def get_queryset(self):
        detail = TransLogDetail.objects.select_related()
        log = TransLog.objects.select_related('constn_site','turn_site','carinfo').filter(turn_site__isnull =False)
        material = Materials.objects

        begin, end = self.get_month_range()
        code = self.request.GET.get("code")

        constn_id = self.request.GET.get("constn_id")
        constn_name = self.request.GET.get("constn_name")

        car_firm = self.request.GET.get("car_firm")
        car_number = self.request.GET.get("car_number")

        matinfo_core = self.request.GET.get("matinfo_core")
        matinfo_name = self.request.GET.get("matinfo_name")
        matinfo_cat = self.request.GET.get("matinfo_cat")
        tran_type = self.request.GET.get("tran_type")

        # 过滤 TransportLog
        if begin and end:
            log = log.filter(build_date__range=[begin, end])
        if code:
            log = log.filter(code__istartswith=code)
        if tran_type:
            log = log.filter(transaction_type=tran_type)
        if constn_id:
            log = log.filter(constn_site__id=constn_id)
        if constn_name:
            log = log.filter(constn_site__name__istartswith=constn_name)
        if car_firm:
            log = log.filter(carinfo__firm__istartswith=car_firm)
        if car_number:
            log = log.filter(carinfo__car_number__istartswith=car_number)

        # 过滤 Materials
        if matinfo_core:
            material = material.filter(mat_code=matinfo_core)
        if matinfo_cat:
            material = material.filter(category=MatCat.objects.get(name=matinfo_cat))
        if matinfo_name:
            material = material.filter(name__istartswith=matinfo_name)

        # 使用过滤后的 log 过滤 detail
        detail = detail.filter(material__in=material.all(), translog__in=log.all())

        return detail.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "出入總表"
        context["whses"] = SiteInfo.objects.filter(genre=0).all()
        context["matinfo_cats"] = MatCat.objects.all()

        return context


class ImportTransportView(ImportDataGeneric):
    title = "上傳EXCEL"
    action = "/transport_log/uploadexcel/"
    html_path = "trans/tramsport_excel_upload.html"
    min_row = 5
    columns3 = [
        "工地",
        "材料規格",
        "倉庫",
        "月報表",
        "施工層別",
        "吊卡車公司",
        "經手人",
    ]

    columns4 = [
        "日期",     "工地編號", "轉單工地",     "業主",
        "工地名稱", "單據編號", "產品編號",     "品名",
        "單位(米)", "米數規格", "入庫量",       "出庫量",
        "入米數",   "出米數",   "入出庫淨額",   "總米數淨額",
        "入料",     "出料",     "數量",         "備註", "層碼",
        "層數",     "公司",     "車號",         "編號", "姓名",
        "key單日期",
    ]

    def upload_excel(self,excel_file):
        workbook = load_workbook(excel_file,read_only=False ,data_only=True)
        worksheet = workbook.get_sheet_by_name("總表")
        if self.check_column( worksheet):
            self.insertDB(worksheet.iter_rows(min_row=self.min_row, values_only=True))
            self.response_data["success"] = True
            self.response_data["msg"] =  "成功"
        else:
            self.response_data["msg"] =  "文件格式不正确"
        """
        _summary_
        後續需要執行的項目
        1. 分割總表與租賃表
        2. 分別分別建立兩個table
        3. 在兩個table中建構create 的function的規則
        4. 再分別建構細項的detial 

        """
    def insertDB(self, data):
        for item in data:
            trancode = excel_value_to_str(item[6])
            mat_code = excel_value_to_str(item[8])
            if trancode is None :
                break
            try:
                if mat_code is None:
                    continue

                remark =  excel_value_to_str(item[20])
                if remark is not None and  "作廢" in remark :
                    TransLogDetail.rollback(trancode)
                    continue 

                tran = TransLog.create(code=trancode,item=item)

                if item[5] is not None and remark is not None and '採購' in remark:
                    DoneSteelReport.add_stock(tran,item) 
                    TransLogDetail.create(tran,item)
                elif item[5] is not None :
                    TransLogDetail.create(tran,item)
                
            except Exception as e:
                # 处理可能的异常情况
                # print(f"{trancode},is error {e} item{item}")
                item_str = ','.join(map(str, item)) + str(e)
                self.error_list.append(item_str)
                logger.info(item_str)
                sys.exit(0)
                

    def check_column(self, worksheet) :
        # 检查列名是否符合预期
        actual_columns3 = self.clean_column_name(worksheet[3])
        actual_columns4 = self.clean_column_name(worksheet[4])
        print(f"actual_columns3:{actual_columns3}")
        print(f"self.columns3:{self.columns3}")
        print(f"actual_columns4:{actual_columns4}")
        print(f"self.columns4:{self.columns4}")
        
        return self.compare_lists(actual_columns3,self.columns3) and self.compare_lists(actual_columns4,self.columns4) 
    

class TransTurn(PageListView):
    model = TransLogDetail
    template_name = "trans/trans_turn.html"
    title_name = "轉單表"

    def get_queryset(self):
        detail = TransLogDetail.objects.select_related()
        log = TransLog.objects.select_related('constn_site','turn_site','carinfo').filter(turn_site__isnull =False).filter(transaction_type="IN")
        material = Materials.objects

        begin, end = self.get_month_range()
        code = self.request.GET.get("code")
        from_site_id = self.request.GET.get("from_constn_id")
        from_site_owner = self.request.GET.get("from_constn_owner")
        from_site_name = self.request.GET.get("from_constn_name")
        to_site_id = self.request.GET.get("to_constn_id")
        to_site_owner = self.request.GET.get("to_constn_owner")
        to_site_name = self.request.GET.get("to_constn_name")

        matinfo_core = self.request.GET.get("matinfo_core")
        matinfo_name = self.request.GET.get("matinfo_name")
        matinfo_cat = self.request.GET.get("matinfo_cat")

        # 过滤 TransportLog
        if begin and end:
            log = log.filter(build_date__range=[begin, end])
        if code:
            log = log.filter(code__istartswith=code)

        if from_site_id:
            log = log.filter(constn_site__code=from_site_id)
        if from_site_owner:
            log = log.filter(constn_site__owner__istartswith=from_site_owner)
        if from_site_name:
            log = log.filter(constn_site__name__istartswith=from_site_name)

        if to_site_id:
            log = log.filter(turn_site__code=to_site_id)
        if to_site_owner:
            log = log.filter(turn_site__owner__istartswith=to_site_owner)
        if to_site_name:
            log = log.filter(turn_site__name__istartswith=to_site_name)

        print(log.query)
        # 过滤 Materials
        if matinfo_core:
            material = material.filter(mat_code=matinfo_core)
        if matinfo_cat:
            material = material.filter(category=MatCat.objects.get(name=matinfo_cat))
        if matinfo_name:
            material = material.filter(name__istartswith=matinfo_name)


        # 使用过滤后的 log 过滤 detail
        detail = detail.filter(material__in=material.all(), translog__in=log.all())

        return detail.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title_name
        context["matinfo_cats"] = MatCat.objects.all()

        return context