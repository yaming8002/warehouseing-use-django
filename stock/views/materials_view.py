import logging
import os
from urllib.parse import quote
from django.core import serializers
from django.http import HttpResponse, JsonResponse
from django.views import View
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

from stock.froms.material import MaterialsForm
from stock.models.material_model import MatCat, Materials, MatSpec
from warehousing_server import settings
from wcom.utils.excel_tool import ImportData2Generic
from wcom.utils.pagelist import PageListView
from wcom.utils.save_control import SaveControlView
from wcom.utils.uitls import excel_value_to_str

logger = logging.getLogger(__name__)


# Create your views here.
class MaterialsView(PageListView):
    model = Materials
    template_name = "stock/materials.html"
    title_name = "物料清單"

    def get_queryset(self):
        result = Materials.objects
        code = self.request.GET.get("mat_code")
        name = self.request.GET.get("name")
        category_id = self.request.GET.get("category_id")
        is_detail = self.request.GET.get("is_detail")

        if name:
            result = result.filter(name__istartswith=name)
        if code:
            result = result.filter(mat_code=code)
        if category_id:
            category = MatCat.objects.get(id=category_id)
            result = result.filter(category=category)
        if not is_detail:
            result = result.filter(specification_id__in =  [23,24])


        return result.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 将查询条件传递到模板
        context["mat_code"] = self.request.GET.get("mat_code", "")
        context["name"] = self.request.GET.get("name", "")
        context["category"] = self.request.GET.get("category_id", "")
        context["category_list"] = MatCat.objects.all()

        return context


class ImportMaterialView(ImportData2Generic):
    title = "上傳EXCEL"
    action = "/material/uploadexcel/"
    columns = [
        "序號",
        "料號",
        "入料號",
        "出料號",
        "料名",
        "分類",
        "規格",
        "耗材",
        "拆分",
        "拆分單位"
    ]

    def insertDB(self, item):
        if item[1] is None :
            return

        id = item[0] if item[0] else None

        material_defaults = {
            'mat_code': excel_value_to_str(item[1]),
            'mat_code2': excel_value_to_str(item[2]),
            'mat_code3': excel_value_to_str(item[3]),
            'name': str(item[4]),
            'category': MatCat.objects.filter(name=item[5]).first(),
            'specification': MatSpec.objects.filter(name=item[6]).first(),
            'is_consumable': item[7] == "是",
            'is_divisible': item[8] == "是",
            'unit_of_division': item[9],
        }

        Materials.objects.update_or_create(
            id=id,
            defaults=material_defaults
        )



class DownloadMaterialView(View):
    def get(self, request, *args, **kwargs):
        filepath = os.path.join(settings.BASE_DIR, "static", "物料品項_模板.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active

        text_style = NamedStyle(name="text_format", number_format="@")
        if "text_format" not in wb.named_styles:
            wb.add_named_style(text_style)

        # 建立資料驗證規則
        dv_category = DataValidation(type="list", formula1="參考!$B$3:$B$6")
        dv_specification = DataValidation(type="list", formula1="參考!$E$3:$E$30")
        dv_yes_no = DataValidation(type="list", formula1="參考!$G$2:$G$3")
        ws.add_data_validation(dv_category)
        ws.add_data_validation(dv_specification)
        ws.add_data_validation(dv_yes_no)
        dv_category.add("F2:F10000")
        dv_specification.add("G2:G10000")
        dv_yes_no.add("H2:H10000")
        dv_yes_no.add("I2:I10000")

        materials = (
            Materials.objects.select_related("category", "specification")
            .all()
            .values_list(
                "id",
                "mat_code",
                "mat_code2",
                "mat_code3",
                "name",
                "is_consumable",
                "is_divisible",
                "unit_of_division",
                "category__name",
                "specification__name",
                named=True,
            )
        )

        for idx, mat in enumerate(materials, start=2):
            ws.cell(row=idx, column=1, value=mat.id)
            ws.cell(row=idx, column=2, value=mat.mat_code).style = text_style
            ws.cell(row=idx, column=3, value=mat.mat_code2).style = text_style
            ws.cell(row=idx, column=4, value=mat.mat_code3).style = text_style
            ws.cell(row=idx, column=5, value=mat.name).style = text_style
            ws.cell(row=idx, column=6, value=mat.category__name)
            ws.cell(row=idx, column=7, value=mat.specification__name)
            ws.cell(row=idx, column=8, value="是" if mat.is_consumable else "否")
            ws.cell(row=idx, column=9, value="是" if mat.is_divisible else "否")
            ws.cell(row=idx, column=10, value=mat.unit_of_division).style = text_style

        filename = "物料清單.xlsx"
        filename_encoded = quote(filename)  # 對文件名進行百分比編碼
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{filename_encoded}"
        )
        wb.save(response)
        return response


class MaterialSeveView(SaveControlView):
    name = "物料"
    model = Materials
    form_class = MaterialsForm


def getMatrtialData(request):
    if request.method == "GET":
        context = {}
        context["matrtials"] = serializers.serialize("json", Materials.objects.all())
        context["matcats"] = serializers.serialize("json", MatCat.objects.all())
        context["spec"] = serializers.serialize("json", MatSpec.objects.all())
        context["success"] = True
        return JsonResponse(context)
