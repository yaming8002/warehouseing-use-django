from datetime import datetime
import os
from django.db.models import Q
from django.conf import settings
from django.core.cache import cache
from openpyxl import load_workbook
from stock.models.material_model import Materials
from stock.models.site_model import SiteInfo
from stock.service.steel_diff_summary import build_constn_diff_view
from trans.models.trans_model import TransLog
from wcom.utils.uitls import to_taiwan_date_format


def get_global_tool_list(update=False):
    setting = cache.get("global_tool_list")
    if update or not setting:
        # 當緩存中沒有資料時，從資料庫中撈取
        setting = (
            Materials.objects.values("id", "name")
            .filter(tool_report=True)
            .order_by("name", "mat_code")
        )
        # 將撈取到的資料存入緩存，並設置過期時間
        cache.set("global_tool_list", setting, timeout=7 * 24 * 60 * 60)  # 七天過期
    return setting


def get_global_component_list(update=False):
    setting = cache.get("global_component_list")
    if update or not setting:
        # 當緩存中沒有資料時，從資料庫中撈取
        setting = (
            Materials.objects.values("id", "name", "component")
            .filter(component__gt=0)
            .order_by("component", "name")
        )
        # 將撈取到的資料存入緩存，並設置過期時間
        cache.set(
            "global_component_list", setting, timeout=7 * 24 * 60 * 60
        )  # 七天過期
    return setting


def get_global_site_json(update=False):
    sitelist = cache.get("global_site_json")
    if update or not sitelist:
        # 當緩存中沒有資料時，從資料庫中撈取
        sitelist = list(
            SiteInfo.objects.exclude(
                Q(name='None') | Q(name='') | Q(name__isnull=True),  # 排除 name 為 'None'、空字串或 None
                Q(owner='None') | Q(owner='') | Q(owner__isnull=True)
              )  # 排除 owner 為 'None'、空字串或 None)
            .values("code", "name", "owner")
            .order_by("code")
            .all()
        )
        print(  SiteInfo.objects.exclude(name="None", owner="None")
            .filter( genre__in=[1, 2] )
            .values("code", "name", "owner")
            .order_by("code")
            .all().query)
        # 將撈取到的資料存入緩存，並設置過期時間
        cache.set("global_site_json", sitelist, timeout=7 * 24 * 60 * 60)  # 七天過期
    return sitelist


def fill_diff_table_excel(constn, output_path):
    # 差異表的EXCEL 匯出
    steel_table, components = build_constn_diff_view(constn)

    # Define the path to the template and output
    template_path = os.path.join(settings.BASE_DIR, "static", "差異表_模板.xlsx")
    workbook = load_workbook(template_path)
    sheet = workbook.active
    first_record = (
        TransLog.objects.values("build_date")
        .filter(constn_site=constn)
        .order_by("build_date")
        .first()
    )
    last_record = (
        TransLog.objects.values("build_date")
        .filter(constn_site=constn)
        .order_by("build_date")
        .last()
    )
    map = {
        "date": to_taiwan_date_format(datetime.now()),
        "begin": to_taiwan_date_format(first_record["build_date"]),
        "end": to_taiwan_date_format(last_record["build_date"]),
        "code": constn.code,
        "owner": constn.owner,
        "name": constn.name,
    }
    replace_cell_value(sheet, map, row_size=5)

    current_row = 9
    for report in steel_table:
        # Insert data for each report into the corresponding columns
        sheet[f"I{current_row}"] = report["input"]["quantity"]
        sheet[f"J{current_row}"] = report["input"]["unit"]
        sheet[f"K{current_row}"] = report["output"]["quantity"]
        sheet[f"L{current_row}"] = report["output"]["unit"]
        sheet[f"M{current_row}"] = report["ng_value"]
        current_row += 1
    # Assuming the first row is the header, start adding data from the second row
    current_row += 2  # Start after the header
    for report in components:
        # Insert data for each report into the corresponding columns
        sheet[f"I{current_row}"] = report["input"]["quantity"]
        sheet[f"K{current_row}"] = report["output"]["quantity"]
        current_row += 1

    # Save the updated Excel file
    workbook.save(output_path)


def replace_cell_value(sheet, map: dict, row_size=5):
    currect_row = 0
    for row in sheet.iter_rows():
        if currect_row > row_size:
            break
        for cell in row:
            for x in map.keys():
                target = "{" + x + "}"
                if cell.value and isinstance(cell.value, str) and target in cell.value:
                    cell.value = cell.value.replace(target, map[x])

        currect_row += 1
