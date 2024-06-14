import json
import logging
# # Create your models here.
import logging.config

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

logging.config.dictConfig(settings.LOGGING)
logger = logging.getLogger(__name__)


class ImportDataGeneric(View):
    title = ""
    action = ""
    min_row = 5
    table_name = ""
    columns = []
    response_data = {"success": False, "msg": "", "error_list": []}
    error_list = []
    html_path = "base/import_data.html"

    def post(self, request, *args, **kwargs):
        try:
            excel_file = request.FILES.get("excel_file")
            if excel_file:
                self.upload_excel(excel_file)
            else:
                self.response_data["msg"] = "未上传 Excel 文件"
        except Exception as e:
            print(e)
            self.response_data["msg"] = str(e)

        cleaned_error_list = []
        for error_item in self.error_list:
            if not isinstance(error_item, ObjectDoesNotExist):
                cleaned_error_list.append(error_item)
        self.response_data["error_list"] = self.error_list

        return JsonResponse(self.response_data)

    def upload_excel(self, excel_file):
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        worksheet = workbook.active
        if self.check_column(worksheet):
            self.insertDB(worksheet.iter_rows(min_row=self.min_row, values_only=True))
            self.response_data["success"] = True
            self.response_data["msg"] = "成功"
        else:
            self.response_data["msg"] = "文件格式不正确"

    def check_column(self, worksheet):
        # 检查列名是否符合预期
        actual_columns = self.clean_column_name(worksheet[1])
        return self.compare_lists(actual_columns, self.columns)

    def clean_column_name(self, sheet: Worksheet):
        lst = []
        for cell in sheet:
            if (
                cell is not None
                and cell.value is not None
                and cell.value.replace(" ", "") != ""
            ):
                lst.append(cell.value.replace(" ", ""))
        return lst

    def insertDB(self, data):
        raise NotImplementedError("Subclasses must implement insertDB method")

    def compare_lists(self, list1, list2):
        # 去除列表中的空白字符
        list1 = [s.strip() for s in list1 if s and s.strip()]
        list2 = [s.strip() for s in list2 if s and s.strip()]

        # 检查两个列表是否相等
        if sorted(list1) == sorted(list2):
            return True
        else:
            return False

    def get(self, request, *args, **kwargs):
        context = {
            "title": self.title,
            "action": self.action,
        }

        return render(request, self.html_path, context)


class ImportData2Generic(View):
    title = "上傳EXCEL"
    action = ""
    min_row = 5
    table_name = ""
    columns = []
    response_data = {"success": False, "msg": "", "error_list": []}
    error_list = []
    html_path = "base/import_data.html"

    def post(self, request, *args, **kwargs):
        jsonData = json.loads(request.body.decode("utf-8"))
        # 检查jsonData是否为列表
        if not isinstance(jsonData, list):
            return JsonResponse(
                {
                    "success": False,
                    "msg": "数据格式错误，期望一个列表",
                    "error_list": [],
                }
            )
        # 处理数据
        error_list = []
        for item in jsonData:
            try:
                if item:
                    self.insertDB(item)
            except Exception as e:
                # 捕获处理数据时的错误，并添加到错误列表中
                errordct = {"item": item, "e": str(e)}
                error_list.append(errordct)
                logger.info(errordct)

        response_data = {
            "success": False if error_list else True,
            "msg": "",
            "error_list": error_list,
        }
        return JsonResponse(response_data)

    def insertDB(self, item):
        pass

    def get(self, request, *args, **kwargs):
        context = {
            "title": self.title,
            "action": self.action,
            "column_name": self.columns,
        }

        return render(request, self.html_path, context)
