from openpyxl import load_workbook
from django.shortcuts import render
from django.http import JsonResponse
from django.views import View

from django.core.exceptions import ObjectDoesNotExist

class ImportDataGeneric(View):
    title = ""
    action = ""
    columns = []
    response_data ={"success":False,"msg":"","error_list":[]}
    error_list = []

    def post(self, request, *args, **kwargs):
        try:
            excel_file = request.FILES.get("excel_file")
            if excel_file:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)
                worksheet = workbook.active
                # 检查列名是否符合预期
                actual_columns = [cell.value for cell in worksheet[1]]
                print(actual_columns)
                print(f"actual_columns:{actual_columns}")
                print(f"self.columns:{self.columns}")
                if actual_columns == self.columns:
                    self.insertDB(worksheet.iter_rows(min_row=2, values_only=True))
                    self.response_data["success"] = True
                    self.response_data["msg"] =  "成功"
                else:
                    self.response_data["msg"] =  "文件格式不正确"
            else:
                self.response_data["msg"] =  "未上传 Excel 文件"
        except Exception as e:
            print(e)
            self.response_data["msg"] =  str(e)

        cleaned_error_list = []
        for error_item in self.error_list:
            if not isinstance(error_item, ObjectDoesNotExist):
                cleaned_error_list.append(error_item)
        self.response_data['error_list'] = self.error_list

        return JsonResponse( self.response_data)

    def insertDB(self, actual_columns):
        raise NotImplementedError("Subclasses must implement insertDB method")

    def get(self, request, *args, **kwargs):
        
        context = {
            "title": self.title,
            "action": self.action,
        }
        return render(request, "base/import_data.html", context)
